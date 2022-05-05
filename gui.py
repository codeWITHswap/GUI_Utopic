from openpyxl import *
from tkinter import *
from functools import partial

path = "Sample.xlsx"
wb = load_workbook(path) 

def do_nothing():
	pass

def kill():
	root1.destroy()
	root2.destroy()
	root3.destroy()
	root4.destroy()
	root5.destroy()

def clear1(var1): 
	sheet = wb["Deposition"]
	for i in range(1,sheet.max_column+1):
		var1[i].delete(0, END)

def clear2(var2):
	sheet = wb["Drying"]
	for i in range(1,sheet.max_column+1):
		var2[i].delete(0, END)

def clear3(var3):
	sheet = wb["Taping"]
	for i in range(1,sheet.max_column+1):
		var3[i].delete(0, END)

def clear4(var4):
	sheet = wb["Lamination"]
	for i in range(1,sheet.max_column+1):
		var4[i].delete(0, END)

def clear5(var5):
	sheet = wb["Running"]
	for i in range(1,sheet.max_column+1):
		var5[i].delete(0, END)

def insert_deposition(var1): 

	sheet = wb["Deposition"]
	current_row = sheet.max_row 

	for i in range(1,sheet.max_column+1):
		if(var1[i].get()==""):
			print("empty sheet")
		else: 
			sheet.cell(row = current_row + 1, column=i).value = var1[i].get()
	# save the file 
	wb.save('Sample.xlsx') 

	var1[1].focus_set() 

	clear1(var1) 

def insert_drying(var2):

	sheet = wb["Drying"]
	current_row = sheet.max_row

	for i in range(1,sheet.max_column+1):
		if(var2[i].get()==""):
			print("empty sheet")
		else:
			sheet.cell(row = current_row + 1, column=i).value = var2[i].get()
	#save the file
	wb.save('Sample.xlsx')
	var2[1].focus_set()

	clear2(var2)

def insert_taping(var3):

	sheet = wb["Taping"]
	current_row = sheet.max_row

	for i in range(1,sheet.max_column+1):
		if(var3[i].get()==""):
			print("empty sheet")
		else:
			sheet.cell(row = current_row + 1, column=i).value = var3[i].get()
	#save the file
	wb.save('Sample.xlsx')
	var3[1].focus_set()

	clear3(var3)

def insert_lamination(var4):

	sheet = wb["Lamination"]
	current_row = sheet.max_row

	for i in range(1,sheet.max_column+1):
		if(var4[i].get()==""):
			print("empty sheet")
		else:
			sheet.cell(row = current_row + 1, column=i).value = var4[i].get()
	#save the file
	wb.save('Sample.xlsx')
	var4[1].focus_set()

	clear4(var4)

def insert_running(var5):

	sheet = wb["Running"]
	current_row = sheet.max_row

	for i in range(1,sheet.max_column+1):
		if(var5[i].get()==""):
			print("empty sheet")
		else:
			sheet.cell(row = current_row + 1, column=i).value = var5[i].get()
	#save the file
	wb.save('Sample.xlsx')
	var5[1].focus_set()

	clear5(var5)
			
def create_deposition_interface():

	root1.deiconify()
	root2.withdraw()
	root3.withdraw()
	root4.withdraw()
	root5.withdraw()
	
	root1.configure(background='light blue') 
	root1.title("Deposition") 
	root1.attributes('-fullscreen',True)

	sheet = wb["Deposition"]
	list1 = [0]
	var1 = [0]

	for i in range(1,sheet.max_column+1):
		if not sheet.cell(row = 1, column = i).value == "": 
			list1.append(sheet.cell(row = 1, column = i).value)

	heading1 = Label(root1, text="Deposition", bg="light blue") 

	for i in range(1,sheet.max_column+1):
		var1.append(Label(root1,text=list1[i],bg="light blue"))
	
	heading1.grid(row=0,column=1)
	for i in range(1,sheet.max_column+1):    
		var1[i].grid(row=i,column=0)

	var1.clear()
	var1.append(0)

	for i in range(1,sheet.max_column+1):
		var1.append(Entry(root1))

	for i in range(1,sheet.max_column):
		var1[i].bind("<Return>",var1[i+1].focus_set())

	for i in range(1,sheet.max_column+1):
		var1[i].grid(row=i, column=1, ipadx="20")  


	insert_deposition_args = partial(insert_deposition,var1)
	submit1 = Button(root1, text="Submit", fg="Black", bg="Red", command=insert_deposition_args) 
	submit1.grid(row=10, column=1,ipadx="50")

	deposition1= Button(root1, text="Deposition", fg="Black", bg="Yellow", command=do_nothing)
	deposition1.grid(row=11, column=0,ipadx="50")

	drying1= Button(root1, text="Drying", fg="Black", bg="Red", command=create_drying_interface)
	drying1.grid(row=11, column=1,ipadx="70")

	taping1= Button(root1, text="Taping", fg="Black", bg="Red", command=create_taping_interface)
	taping1.grid(row=11, column=2,ipadx="50")

	lamination1= Button(root1, text="Lamination", fg="Black", bg="Red", command=create_lamination_interface)
	lamination1.grid(row=11, column=3,ipadx="50")

	running1= Button(root1, text="Running", fg="Black", bg="Red", command=create_running_interface)
	running1.grid(row=11, column=4,ipadx="50")

	exit1 = Button(root1, text="Exit", fg="Black", bg="Red", command=kill)
	exit1.grid(row=10, column=2,ipadx="50")

def create_drying_interface():

	root1.withdraw()
	root2.deiconify()
	root3.withdraw()
	root4.withdraw()
	root5.withdraw()

	root2.configure(background='light blue') 
	root2.title("Drying") 
	root2.attributes('-fullscreen',True)

	sheet = wb["Drying"]
	list2 = [0]
	var2 = [0]

	for i in range(1,sheet.max_column+1):
		if not sheet.cell(row = 1, column = i).value == "": 
			list2.append(sheet.cell(row = 1, column = i).value)

	heading2 = Label(root2, text="Drying", bg="light blue") 
	for i in range(1,sheet.max_column+1):
		var2.append(Label(root2,text=list2[i],bg="light blue"))

	heading2.grid(row=0,column=1)
	for i in range(1,sheet.max_column+1):    
		var2[i].grid(row=i,column=0)

	var2.clear()
	var2.append(0)

	for i in range(1,sheet.max_column+1):
		var2.append(Entry(root2))

	for i in range(1,sheet.max_column):
		var2[i].bind("<Return>",var2[i+1].focus_set())

	for i in range(1,sheet.max_column+1):
		var2[i].grid(row=i, column=1, ipadx="20")  


	insert_drying_args = partial(insert_drying,var2)
	submit2 = Button(root2, text="Submit", fg="Black", bg="Red", command=insert_drying_args) 
	submit2.grid(row=10, column=1,ipadx="50")

	deposition2= Button(root2, text="Deposition", fg="Black", bg="Red", command=create_deposition_interface)
	deposition2.grid(row=11, column=0,ipadx="50")

	drying2= Button(root2, text="Drying", fg="Black", bg="Yellow", command=do_nothing)
	drying2.grid(row=11, column=1,ipadx="70")

	taping2= Button(root2, text="Taping", fg="Black", bg="Red", command=create_taping_interface)
	taping2.grid(row=11, column=2,ipadx="50")

	lamination2= Button(root2, text="Lamination", fg="Black", bg="Red", command=create_lamination_interface)
	lamination2.grid(row=11, column=3,ipadx="50")

	running2= Button(root2, text="Running", fg="Black", bg="Red", command=create_running_interface)
	running2.grid(row=11, column=4,ipadx="50")


	exit2 = Button(root2, text="Exit", fg="Black", bg="Red", command=kill)
	exit2.grid(row=10, column=2,ipadx="50")
def create_taping_interface():

	root1.withdraw()
	root2.withdraw()
	root3.deiconify()
	root4.withdraw()
	root4.withdraw()

	root3.configure(background='light blue') 
	root3.title("Taping") 
	root3.attributes('-fullscreen',True)

	sheet = wb["Taping"]
	list3 = [0]
	var3 = [0]

	for i in range(1,sheet.max_column+1):
		if not sheet.cell(row = 1, column = i).value == "": 
			list3.append(sheet.cell(row = 1, column = i).value)

	heading3 = Label(root3, text="Taping", bg="light blue") 
	for i in range(1,sheet.max_column+1):
		var3.append(Label(root3,text=list3[i],bg="light blue"))

	heading3.grid(row=0,column=1)
	for i in range(1,sheet.max_column+1):    
		var3[i].grid(row=i,column=0)

	var3.clear()
	var3.append(0)

	for i in range(1,sheet.max_column+1):
		var3.append(Entry(root3))

	for i in range(1,sheet.max_column):
		var3[i].bind("<Return>",var3[i+1].focus_set())

	for i in range(1,sheet.max_column+1):
		var3[i].grid(row=i, column=1, ipadx="20")  


	insert_taping_args = partial(insert_taping,var3)
	submit3 = Button(root3, text="Submit", fg="Black", bg="Red", command=insert_taping_args) 
	submit3.grid(row=10, column=1,ipadx="50")

	deposition3= Button(root3, text="Deposition", fg="Black", bg="Red", command=create_deposition_interface)
	deposition3.grid(row=11, column=0,ipadx="50")

	drying3= Button(root3, text="Drying", fg="Black", bg="Red", command=create_drying_interface)
	drying3.grid(row=11, column=1,ipadx="70")

	taping3= Button(root3, text="Taping", fg="Black", bg="Yellow", command=do_nothing)
	taping3.grid(row=11, column=2,ipadx="50")

	lamination3= Button(root3, text="Lamination", fg="Black", bg="Red", command=create_lamination_interface)
	lamination3.grid(row=11, column=3,ipadx="50")

	running3= Button(root3, text="Running", fg="Black", bg="Red", command=create_running_interface)
	running3.grid(row=11, column=4,ipadx="50")

	exit3 = Button(root3, text="Exit", fg="Black", bg="Red", command=kill)
	exit3.grid(row=10, column=2,ipadx="50")

def create_lamination_interface():

	root1.withdraw()
	root2.withdraw()
	root3.withdraw()
	root4.deiconify()
	root5.withdraw()

	root4.configure(background='light blue') 
	root4.title("Lamination") 
	root4.attributes('-fullscreen',True)

	sheet = wb["Lamination"]
	list4 = [0]
	var4 = [0]

	for i in range(1,sheet.max_column+1):
		if not sheet.cell(row = 1, column = i).value == "": 
			list4.append(sheet.cell(row = 1, column = i).value)

	heading4 = Label(root4, text="Lamination", bg="light blue") 
	for i in range(1,sheet.max_column+1):
		var4.append(Label(root4,text=list4[i],bg="light blue"))

	heading4.grid(row=0,column=1)
	for i in range(1,sheet.max_column+1):    
		var4[i].grid(row=i,column=0)

	var4.clear()
	var4.append(0)

	for i in range(1,sheet.max_column+1):
		var4.append(Entry(root4))

	for i in range(1,sheet.max_column):
		var4[i].bind("<Return>",var4[i+1].focus_set())

	for i in range(1,sheet.max_column+1):
		var4[i].grid(row=i, column=1, ipadx="20")  


	insert_lamination_args = partial(insert_lamination,var4)
	submit4 = Button(root4, text="Submit", fg="Black", bg="Red", command=insert_lamination_args) 
	submit4.grid(row=10, column=1,ipadx="50")

	deposition4= Button(root4, text="Deposition", fg="Black", bg="Red", command=create_deposition_interface)
	deposition4.grid(row=11, column=0,ipadx="50")

	drying4= Button(root4, text="Drying", fg="Black", bg="Red", command=create_drying_interface)
	drying4.grid(row=11, column=1,ipadx="70")

	taping4= Button(root4, text="Taping", fg="Black", bg="Red", command=create_taping_interface)
	taping4.grid(row=11, column=2,ipadx="50")

	lamination4= Button(root4, text="Lamination", fg="Black", bg="Yellow", command=do_nothing)
	lamination4.grid(row=11, column=3,ipadx="50")

	running4= Button(root4, text="Running", fg="Black", bg="Red", command=create_running_interface)
	running4.grid(row=11, column=4,ipadx="50")

	exit4 = Button(root4, text="Exit", fg="Black", bg="Red", command=kill)
	exit4.grid(row=10, column=2,ipadx="50")

def create_running_interface():

	root1.withdraw()
	root2.withdraw()
	root3.withdraw()
	root4.withdraw()
	root5.deiconify()

	root5.configure(background='light blue') 
	root5.title("Running") 
	root5.attributes('-fullscreen',True)

	sheet = wb["Running"]
	list5 = [0]
	var5 = [0]

	for i in range(1,sheet.max_column+1):
		if not sheet.cell(row = 1, column = i).value == "": 
			list5.append(sheet.cell(row = 1, column = i).value)

	heading5 = Label(root5, text="Running", bg="light blue") 
	for i in range(1,sheet.max_column+1):
		var5.append(Label(root5,text=list5[i],bg="light blue"))

	heading5.grid(row=0,column=1)
	for i in range(1,sheet.max_column+1):    
		var5[i].grid(row=i,column=0)

	var5.clear()
	var5.append(0)

	for i in range(1,sheet.max_column+1):
		var5.append(Entry(root5))

	for i in range(1,sheet.max_column):
		var5[i].bind("<Return>",var5[i+1].focus_set())

	for i in range(1,sheet.max_column+1):
		var5[i].grid(row=i, column=1, ipadx="20")  


	insert_running_args = partial(insert_running,var5)
	submit5 = Button(root5, text="Submit", fg="Black", bg="Red", command=insert_running_args) 
	submit5.grid(row=10, column=1,ipadx="50")

	deposition5= Button(root5, text="Deposition", fg="Black", bg="Red", command=create_deposition_interface)
	deposition5.grid(row=11, column=0,ipadx="50")

	drying5= Button(root5, text="Drying", fg="Black", bg="Red", command=create_drying_interface)
	drying5.grid(row=11, column=1,ipadx="70")

	taping5= Button(root5, text="Taping", fg="Black", bg="Red", command=create_taping_interface)
	taping5.grid(row=11, column=2,ipadx="50")

	lamination5= Button(root5, text="Lamination", fg="Black", bg="Red", command=create_lamination_interface)
	lamination5.grid(row=11, column=3,ipadx="50")

	running5= Button(root5, text="Running", fg="Black", bg="Yellow", command=do_nothing)
	running5.grid(row=11, column=4,ipadx="50")

	exit5 = Button(root5, text="Exit", fg="Black", bg="Red", command=kill)
	exit5.grid(row=10, column=2,ipadx="50")

# Driver code 
if __name__ == "__main__": 
	
	root1=Tk()	# For Deposition
	root2=Tk()	# For Drying
	root3=Tk()	# For Taping
	root4=Tk()	# For Lamination
	root5=Tk()	# For Running

	root2.withdraw()
	root3.withdraw()
	root4.withdraw()
	root5.withdraw()

	create_deposition_interface()	

	# start the GUI 
	root1.mainloop()
